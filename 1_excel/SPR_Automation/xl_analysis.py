from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from copy import copy, deepcopy
import sys
import re

def get_position (base_position, index):
    digit =  re.findall("[0-9]+", base_position)
    character = re.findall("[A-Z]+", base_position)

    ascii_code = ord(character[0])

    new_ascii_code = ascii_code + 2 * index

    # 0x41 means "A" in ascii code.
    dividen = int((new_ascii_code - 0X41) / 0X1A)  # 0x1B meadn total# of capital
    modulus = (new_ascii_code - 0X41) % 0X1A

    new_ascii_code = 0x41 + modulus
    new_character = chr(new_ascii_code)

    # print("jinha", dividen, modulus, new_character)

    if (dividen): 
        addedcharcode = 0x41 + (dividen - 1)   #make AA. AC if index is beyond Z  
        appended_character = chr(addedcharcode)
        new_position = appended_character[0] + new_character[0] + digit[0]
    else:
        new_position = new_character[0] + digit[0]

    # print("jinha2", new_position)
    return new_position

def get_delta_position (base_position):

    digit =  re.findall("[0-9]+", base_position)
    character = re.findall("[A-Z]+", base_position)

    # print ("jinha5", character[0], digit )

    if(len(character[0]) > 1 ):
        ascii_code = ord(character[0][1:])
        appendedchar = character[0][:1]  # In case AA, AB, A 
        # print ("jinha7", appendedchar, ascii_code )
    else:
        ascii_code = ord(character[0])

    delta_ascii_code =  ascii_code - 1  #  delta position is the previous one 
    prev_ascii_code =  ascii_code  - 2   #  pre position is the two column back

    if((len(character[0]) > 1) and (prev_ascii_code < 0x41) ):
        delta_character = chr(delta_ascii_code)
        prev_character = chr(prev_ascii_code + 0x1A)
        delta_position = appendedchar + delta_character[0] + digit[0]
        prev_position =  prev_character[0] + digit[0]
    elif((len(character[0]) > 1) and (prev_ascii_code >= 0x41) ):
        delta_character = chr(delta_ascii_code)
        prev_character = chr(prev_ascii_code)
        delta_position = appendedchar + delta_character[0] + digit[0]
        prev_position = appendedchar + prev_character[0] + digit[0]
    else:
        delta_character = chr(delta_ascii_code)
        prev_character = chr(prev_ascii_code)
        delta_position = delta_character[0] + digit[0]
        prev_position =  prev_character[0] + digit[0]

    return [delta_position, prev_position] 

def update_delta(ws_summary, new_position):
    return_position = get_delta_position (new_position)
    txt = "={0}-{1}".format(new_position, return_position[1])
    ws_summary[return_position[0]] = txt


file_name = "osprey_issues_master.xlsx"
txt = ""

total_spr_table = {
    'total' : 'B2',
    'total_NC' : 'B3',
    'total_IO' : 'B4',
    'Open' : 'B5',
    'Open_NC' : 'B6',
    'Open_IO' : 'B7',
    'Resolved' : 'B8',
    'Verified' : 'B9',
    'Closed' : 'B10',
}


sw_spr_table = {
    'SW_total' : 'B15',
    'SW_total_NC' : 'B16',
    'SW_total_IO' : 'B17',
    'SW_Open' : 'B18',
    'SW_Open_NC' : 'B19',
    'SW_Open_IO' : 'B20',
    'SW_Resolved' : 'B21',
    'SW_Verified' : 'B22',
    'SW_Closed' : 'B23',
}


system_spr_table = {
    'system_total' : 'B29',
    'system_total_NC' : 'B30',
    'system_total_IO' : 'B31',
    'system_Open' : 'B32',
    'system_Open_NC' : 'B33',
    'system_Open_IO' : 'B34',
    'system_Resolved' : 'B35',
    'system_Verified' : 'B36',
    'system_Closed' : 'B37',
}

sw_members = [
        "\"=dongyoung.choi\"",
        "\"=taeyang.an\"",
        "\"=han il.lee\"",
        "\"=ho.lee\"",
        "\"=youjin.jung\"",        
        "\"=doo je.sung\"",
        "\"=jin ha.hwang\"",
        "\"=youngdug.kim\"",
        "\"=yujin.na\"",
        "\"=kyudong.kim\"",
        "\"=heejin.na\"",
        "\"=Suncheol.Heo\"",
]

system_members = [
        "\"=jihye.han\"",
        "\"=bong hyo.han\"",
        "\"=jae ha.hyun\"",
        "\"=jong gun.lee\"",
        "\"=jungho.kim\"",
        "\"=taeyun.kim\"",
        "\"=yasuhiro.yamada\"",
        "\"=dongwoo.lee\""
]


if len(sys.argv) == 1:
    txt = "running : Python xl_analysis.py {} ".format(file_name)
    print(txt) 
elif len(sys.argv) == 2:
    file_name = sys.argv[1]
    txt = "running : Python xl_analysis.py {} ".format(file_name)
    print(txt) 
else:  
    print("Usage : Python xl_analysis.py file_name.xlsx")
    sys.exit()

wb_src = load_workbook(file_name)
ws_srcs = []  # using list 
flag_for_analysis = {} # using dictionalry type

print(wb_src.sheetnames)

for sheetname in wb_src.sheetnames:
    validsheet = sheetname.find("FW")
    if(validsheet != -1):  # valid weekly spr worksheet exists
        flag_for_analysis[sheetname] = True
    else:
        flag_for_analysis[sheetname] = False

# getting target worksheet for analysis
ws_summary = wb_src[wb_src.sheetnames[0]]

# adding weekly base SPR worksheet only
for key in flag_for_analysis:
    val = flag_for_analysis[key]
    if (val == True):
        ws_srcs.append(wb_src[key])


print(ws_srcs)
print(ws_summary)

index = 0
ascii = 0
for ws_src in reversed(ws_srcs):

    new_position = get_position (total_spr_table['total'], index)
    # print("jinha4", new_position);
    txt = "=COUNTA('{}'!A2:A3000)".format(ws_src.title)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position)
           

    #NC from ALL SPR
    new_position = get_position (total_spr_table['total_NC'], index)
    txt = "=COUNTIF('{0}'!J2:J3000,\"=NC - Design Non-Conformance\")".format(ws_src.title)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position)

    #IO from ALL SPR
    new_position = get_position (total_spr_table['total_IO'], index)
    txt = "=COUNTIFS('{0}'!J2:J3000,\"=IO - Improvement Opportunity\")".format(ws_src.title)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position)

    # Open from  ALL SPR
    new_position = get_position (total_spr_table['Open'], index)
    txt = "=COUNTIFS('{0}'!M2:M3000,\"=Submitted\")\
    +COUNTIFS('{0}'!M2:M3000,\"=Accepted\")\
    +COUNTIFS('{0}'!M2:M3000,\"=Failed\")\
    +COUNTIFS('{0}'!M2:M3000,\"=In Progress\")\
    +COUNTIFS('{0}'!M2:M3000,\"=In Review\")\
    +COUNTIFS('{0}'!M2:M3000,\"=Postponed\")".format(ws_src.title)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position)

    #NC for Open
    new_position = get_position (total_spr_table['Open_NC'], index)
    txt = "=COUNTIFS('{0}'!M2:M3000,\"=Submitted\",'{0}'!J2:J3000,\"=NC - Design Non-Conformance\")\
    +COUNTIFS('{0}'!M2:M3000,\"=Accepted\",'{0}'!J2:J3000,\"=NC - Design Non-Conformance\")\
    +COUNTIFS('{0}'!M2:M3000,\"=Failed\",'{0}'!J2:J3000,\"=NC - Design Non-Conformance\")\
    +COUNTIFS('{0}'!M2:M3000,\"=In Progress\",'{0}'!J2:J3000,\"=NC - Design Non-Conformance\")\
    +COUNTIFS('{0}'!M2:M3000,\"=In Review\",'{0}'!J2:J3000,\"=NC - Design Non-Conformance\")\
    +COUNTIFS('{0}'!M2:M3000,\"=Postponed\",'{0}'!J2:J3000,\"=NC - Design Non-Conformance\")".format(ws_src.title)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position)


    #IO for Open
    new_position = get_position (total_spr_table['Open_IO'], index)
    txt = "=COUNTIFS('{0}'!M2:M3000,\"=Submitted\",'{0}'!J2:J3000,\"=IO - Improvement Opportunity\")\
    +COUNTIFS('{0}'!M2:M3000,\"=Accepted\",'{0}'!J2:J3000,\"=IO - Improvement Opportunity\")\
    +COUNTIFS('{0}'!M2:M3000,\"=Failed\",'{0}'!J2:J3000,\"=IO - Improvement Opportunity\")\
    +COUNTIFS('{0}'!M2:M3000,\"=In Progress\",'{0}'!J2:J3000,\"=IO - Improvement Opportunity\")\
    +COUNTIFS('{0}'!M2:M3000,\"=In Review\",'{0}'!J2:J3000,\"=IO - Improvement Opportunity\")\
    +COUNTIFS('{0}'!M2:M3000,\"=Postponed\",'{0}'!J2:J3000,\"=IO - Improvement Opportunity\")".format(ws_src.title)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position)

    #Resolved 
    new_position = get_position (total_spr_table['Resolved'], index)
    txt = "=COUNTIFS('{0}'!M2:M3000,\"=Resolved\")".format(ws_src.title)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position)

    #Verified 
    new_position = get_position (total_spr_table['Verified'], index)
    txt = "=COUNTIFS('{0}'!M2:M3000,\"=Verified\")".format(ws_src.title)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position)

    #Closed 
    new_position = get_position (total_spr_table['Closed'], index)
    txt = "=COUNTIFS('{0}'!M2:M3000,\"=Closed\")".format(ws_src.title)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position)


    # # total SPRs owned by s/w team
    temp_txt="="
    for sw_member in sw_members:
        temp_txt += "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ")+"
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (sw_spr_table['SW_total'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position)


    # # NC SPRs owned by s/w team
    temp_txt="="
    for sw_member in sw_members:
        temp_txt += "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\")".format(ws_src.title) +"+"
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (sw_spr_table['SW_total_NC'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position)


    # IO  SPRs owned by s/w team
    temp_txt="="
    for sw_member in sw_members:
        temp_txt += "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!J2:J3000,\"=IO - Improvement Opportunity\")".format(ws_src.title) +"+"
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (sw_spr_table['SW_total_IO'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position)


    # # Open SPRs owned by s/w team
    temp_txt="="
    temp_txt1=""
    temp_txt2=""
    temp_txt3=""
    temp_txt4=""
    temp_txt5=""
    temp_txt6=""
    for sw_member in sw_members:
        temp_txt1 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000,\"=Submitted\")".format(ws_src.title) + "+" 
        temp_txt2 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000,\"=Accepted\")".format(ws_src.title) + "+"
        temp_txt3 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000,\"=Failed\")".format(ws_src.title) + "+"
        temp_txt4 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000,\"=In Review\")".format(ws_src.title) + "+"
        temp_txt5 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000,\"=Postponed\")".format(ws_src.title) + "+"
        temp_txt6 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000,\"=In Progress\")".format(ws_src.title) + "+"
        temp_txt += temp_txt1 + temp_txt2 + temp_txt3 + temp_txt4 + temp_txt5 + temp_txt6
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (sw_spr_table['SW_Open'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position)


 
    # # Open NC SPRs owned by s/w team
    temp_txt="="
    temp_txt1=""
    temp_txt2=""
    temp_txt3=""
    temp_txt4=""
    temp_txt5=""
    temp_txt6=""
    for sw_member in sw_members:
        temp_txt1 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000,\"=Submitted\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\")".format(ws_src.title) + "+" 
        temp_txt2 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000,\"=Accepted\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\")".format(ws_src.title) + "+"
        temp_txt3 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000,\"=Failed\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\")".format(ws_src.title) + "+"
        temp_txt4 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000,\"=In Review\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\")".format(ws_src.title) + "+"
        temp_txt5 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000,\"=Postponed\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\")".format(ws_src.title) + "+"
        temp_txt6 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000,\"=In Progress\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\")".format(ws_src.title) + "+"
        temp_txt += temp_txt1 + temp_txt2 + temp_txt3 + temp_txt4 + temp_txt5 + temp_txt6
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (sw_spr_table['SW_Open_NC'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position)

    # # Open IO SPRs owned by s/w team
    temp_txt="="
    temp_txt1=""
    temp_txt2=""
    temp_txt3=""
    temp_txt4=""
    temp_txt5=""
    temp_txt6=""
    for sw_member in sw_members:
        temp_txt1 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000,\"=Submitted\"".format(ws_src.title) + ",{0}!J2:J3000,\"=IO - Improvement Opportunity\")".format(ws_src.title) + "+" 
        temp_txt2 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000,\"=Accepted\"".format(ws_src.title) + ",{0}!J2:J3000,\"=IO - Improvement Opportunity\")".format(ws_src.title) + "+"
        temp_txt3 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000,\"=Failed\"".format(ws_src.title) + ",{0}!J2:J3000,\"=IO - Improvement Opportunity\")".format(ws_src.title) + "+"
        temp_txt4 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000,\"=In Review\"".format(ws_src.title) + ",{0}!J2:J3000,\"=IO - Improvement Opportunity\")".format(ws_src.title) + "+"
        temp_txt5 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000,\"=Postponed\"".format(ws_src.title) + ",{0}!J2:J3000,\"=IO - Improvement Opportunity\")".format(ws_src.title) + "+"
        temp_txt6 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000,\"=In Progress\"".format(ws_src.title) + ",{0}!J2:J3000,\"=IO - Improvement Opportunity\")".format(ws_src.title) + "+"
        temp_txt += temp_txt1 + temp_txt2 + temp_txt3 + temp_txt4 + temp_txt5 + temp_txt6
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (sw_spr_table['SW_Open_IO'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position)


    # # Resolved SPRs owned by s/w team
    temp_txt="="
    for sw_member in sw_members:
        temp_txt += "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000, \"=Resolved\")".format(ws_src.title) +"+"
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (sw_spr_table['SW_Resolved'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position)

    # # Verified SPRs owned by s/w team
    temp_txt="="
    for sw_member in sw_members:
        temp_txt += "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000, \"=Verified\")".format(ws_src.title) +"+"
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (sw_spr_table['SW_Verified'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position)

    # # Closed SPRs owned by s/w team
    temp_txt="="
    for sw_member in sw_members:
        temp_txt += "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + sw_member + ",{0}!M2:M3000, \"=Closed\")".format(ws_src.title) +"+"
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (sw_spr_table['SW_Closed'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position)

    # # total SPRs owned by system team
    temp_txt="="
    for system_member in system_members:
        temp_txt += "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ")+"
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (system_spr_table['system_total'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position)


    # # NC SPRs owned by system team
    temp_txt="="
    for system_member in system_members:
        temp_txt += "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\")".format(ws_src.title) +"+"
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (system_spr_table['system_total_NC'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position)

    # IO  SPRs owned by system team
    temp_txt="="
    for system_member in system_members:
        temp_txt += "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!J2:J3000,\"=IO - Improvement Opportunity\")".format(ws_src.title) +"+"
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (system_spr_table['system_total_IO'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position)


    # # Open SPRs owned by system team
    temp_txt="="
    temp_txt1=""
    temp_txt2=""
    temp_txt3=""
    temp_txt4=""
    temp_txt5=""
    temp_txt6=""
    for system_member in system_members:
        temp_txt1 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000,\"=Submitted\")".format(ws_src.title) + "+" 
        temp_txt2 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000,\"=Accepted\")".format(ws_src.title) + "+"
        temp_txt3 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000,\"=Failed\")".format(ws_src.title) + "+"
        temp_txt4 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000,\"=In Review\")".format(ws_src.title) + "+"
        temp_txt5 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000,\"=Postponed\")".format(ws_src.title) + "+"
        temp_txt6 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000,\"=In Progress\")".format(ws_src.title) + "+"
        temp_txt += temp_txt1 + temp_txt2 + temp_txt3 + temp_txt4 + temp_txt5 + temp_txt6
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (system_spr_table['system_Open'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position)


 
    # # Not Started NC SPRs owned by system team
    temp_txt="="
    temp_txt1=""
    temp_txt2=""
    temp_txt3=""
    temp_txt4=""
    temp_txt5=""
    temp_txt6=""
    for system_member in system_members:
        temp_txt1 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000,\"=Submitted\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\")".format(ws_src.title) + "+" 
        temp_txt2 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000,\"=Accepted\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\")".format(ws_src.title) + "+"
        temp_txt3 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000,\"=Failed\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\")".format(ws_src.title) + "+"
        temp_txt4 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000,\"=In Review\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\")".format(ws_src.title) + "+"
        temp_txt5 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000,\"=Postponed\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\")".format(ws_src.title) + "+"
        temp_txt6 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000,\"=In Progress\"".format(ws_src.title) + ",{0}!J2:J3000,\"=NC - Design Non-Conformance\")".format(ws_src.title) + "+"
        temp_txt += temp_txt1 + temp_txt2 + temp_txt3 + temp_txt4 + temp_txt5 + temp_txt6
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (system_spr_table['system_Open_NC'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position)


    # # OPen IO SPRs owned by system team
    temp_txt="="
    temp_txt1=""
    temp_txt2=""
    temp_txt3=""
    temp_txt4=""
    temp_txt5=""
    temp_txt6=""
    for system_member in system_members:
        temp_txt1 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000,\"=Submitted\"".format(ws_src.title) + ",{0}!J2:J3000,\"=IO - Improvement Opportunity\")".format(ws_src.title) + "+" 
        temp_txt2 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000,\"=Accepted\"".format(ws_src.title) + ",{0}!J2:J3000,\"=IO - Improvement Opportunity\")".format(ws_src.title) + "+"
        temp_txt3 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000,\"=Failed\"".format(ws_src.title) + ",{0}!J2:J3000,\"=IO - Improvement Opportunity\")".format(ws_src.title) + "+"
        temp_txt4 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000,\"=In Review\"".format(ws_src.title) + ",{0}!J2:J3000,\"=IO - Improvement Opportunity\")".format(ws_src.title) + "+"
        temp_txt5 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000,\"=Postponed\"".format(ws_src.title) + ",{0}!J2:J3000,\"=IO - Improvement Opportunity\")".format(ws_src.title) + "+"
        temp_txt6 = "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000,\"=In Progress\"".format(ws_src.title) + ",{0}!J2:J3000,\"=IO - Improvement Opportunity\")".format(ws_src.title) + "+"
        temp_txt += temp_txt1 + temp_txt2 + temp_txt3 + temp_txt4 + temp_txt5 + temp_txt6
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (system_spr_table['system_Open_IO'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position)

    # # Resolved SPRs owned by system team
    temp_txt="="
    for system_member in system_members:
        temp_txt += "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000, \"=Resolved\")".format(ws_src.title) +"+"
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (system_spr_table['system_Resolved'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position)

    # # Verified SPRs owned by system team
    temp_txt="="
    for system_member in system_members:
        temp_txt += "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000, \"=Verified\")".format(ws_src.title) +"+"
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (system_spr_table['system_Verified'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position)

    # # Closed SPRs owned by system team
    temp_txt="="
    for system_member in system_members:
        temp_txt += "COUNTIFS({0}!N2:N3000,".format(ws_src.title) + system_member + ",{0}!M2:M3000, \"=Closed\")".format(ws_src.title) +"+"
    txt = temp_txt[0:len(temp_txt)-1]  # truncate last "+"
    new_position = get_position (system_spr_table['system_Closed'], index)
    ws_summary[new_position] = txt
    if (index != 0):
        update_delta(ws_summary, new_position)

    index += 1

wb_src.save(file_name)
wb_src.close()
