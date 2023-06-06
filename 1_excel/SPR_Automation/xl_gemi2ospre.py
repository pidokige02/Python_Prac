from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill

from copy import copy, deepcopy
import sys


new_file_name = "GEMINI_R4_FW12.xlsx"
old_file_name = "GEMINI_R4_FW12_SMART.xlsx"

if len(sys.argv) == 1:
    txt = "running : Python xl_gemi2ospre.py {} {}".format(old_file_name, new_file_name)
    print(txt) 
elif len(sys.argv) == 2:
    print("Usage : Python xl_gemi2ospre.py oldfilename newfilename!")
    txt = "running : Python xl_gemi2ospre.py {} {}".format(old_file_name, new_file_name)
    print(txt) 
    sys.exit()
elif len(sys.argv) == 3:
    old_file_name = sys.argv[1]
    new_file_name = sys.argv[2]
    txt = "running : Python xl_gemi2ospre.py {} {}".format(old_file_name, new_file_name)
    print(txt) 
else:  
    print("Usage : Python xl_gemi2ospre.py oldfilename newfilename!")
    sys.exit()

wb_old = load_workbook(old_file_name)
wb_new = load_workbook(new_file_name)

print(old_file_name)
print(new_file_name)


print(wb_new.sheetnames)
print(wb_old.sheetnames)

total_name_table = [
  {
    "RP" : "ajay.s",
    "GEUK" : "dongyoung.choi"
  },
  {
    "RP" : "alan.meyers",
    "GEUK" : "han il.lee"
  },
  {
    "RP" : "anurag.velekkattsunilkumar",
    "GEUK" : "kyudong.kim"
  },
  {
    "RP" : "arun.ravi1",
    "GEUK" : "youngdug.kim"
  },
  {
    "RP" : "arya.ck",
    "GEUK" : "kyudong.kim"
  },
  {
    "RP" : "christin.mathew",
    "GEUK" : "N/A"
  },
  {
    "RP" : "chul wook.moon",
    "GEUK" : "doo je.sung"
  },
  {
    "RP" : "cinu.ts",
    "GEUK" : "han il.lee"
  },
  {
    "RP" : "doo je.sung",
    "GEUK" : "doo je.sung"
  },
  {
    "RP" : "ernest.passaretti",
    "GEUK" : "kyudong.kim"
  },
  {
    "RP" : "frederick.frigo",
    "GEUK" : "youjin.na"
  },
  {
    "RP" : "gopi.pattaswamy",
    "GEUK" : "dongyoung.choi"
  },
  {
    "RP" : "gopi.pattaswamy",
    "GEUK" : "dongyoung.choi"
  },
  {
    "RP" : "han il.lee",
    "GEUK" : "han il.lee"
  },
  {
    "RP" : "hanako.kato",
    "GEUK" : "ho.lee"
  },
  {
    "RP" : "hee dong.lee",
    "GEUK" : "doo je.sung"
  },
  {
    "RP" : "ho.lee",
    "GEUK" : "ho.lee"
  },
  {
    "RP" : "huisu.jeong",
    "GEUK" : "ho.lee"
  },
  {
    "RP" : "james.zhang",
    "GEUK" : "system"
  },
  {
    "RP" : "joohyun.song",
    "GEUK" : "han il.lee"
  },
  {
    "RP" : "komal.dutta",
    "GEUK" : "system"
  },
  {
    "RP" : "krishnamoorthy.a",
    "GEUK" : "dongyoung.choi"
  },
  {
    "RP" : "linta.george",
    "GEUK" : "youngdug.kim"
  },
  {
    "RP" : "liza.anna antony",
    "GEUK" : "TBD"
  },
  {
    "RP" : "mahamadou.diakite",
    "GEUK" : "han il.lee"
  },
  {
    "RP" : "manabu.arima",
    "GEUK" : "TBD"
  },
  {
    "RP" : "manoj.nair",
    "GEUK" : "youngdug.kim"
  },
  {
    "RP" : "masaru.ogasawara",
    "GEUK" : "han il.lee"
  },
  {
    "RP" : "mayumi.ito",
    "GEUK" : "taeyang.an"
  },
  {
    "RP" : "michael.harland",
    "GEUK" : "han il.lee"
  },
  {
    "RP" : "michael.seiler",
    "GEUK" : "youngdug.kim"
  },
  {
    "RP" : "miki.konishi",
    "GEUK" : "taeyang.an"
  },
  {
    "RP" : "nathan.luttmann",
    "GEUK" : "han il.lee"
  },
  {
    "RP" : "nithin.georgekuttyv",
    "GEUK" : "ho.lee"
  },
  {
    "RP" : "paul.odea",
    "GEUK" : "youngdug.kim"
  },
  {
    "RP" : "richard.kulakowski",
    "GEUK" : "TBD"
  },
  {
    "RP" : "sangjae.kim",
    "GEUK" : "sangjae.kim"
  },
  {
    "RP" : "sarath.sathikumar",
    "GEUK" : "TBD"
  },
  {
    "RP" : "satoru.takahashir",
    "GEUK" : "taeyang.an"
  },
  {
    "RP" : "scott.coursin",
    "GEUK" : "taeyang.an"
  },
  {
    "RP" : "seiji.funaya",
    "GEUK" : "dongyoung.choi"
  },
  {
    "RP" : "shiji.mohan",
    "GEUK" : "dongyoung.choi"
  },
  {
    "RP" : "shingo.nishiyama",
    "GEUK" : "doo je.sung"
  },
  {
    "RP" : "syed.ah",
    "GEUK" : "youngdug.kim"
  },
  {
    "RP" : "taeyang.an",
    "GEUK" : "taeyang.an"
  },
  {
    "RP" : "velayudhan.k",
    "GEUK" : "doo je.sung"
  },
  {
    "RP" : "yelena.tsymbalenko",
    "GEUK" : "ho.lee"
  },
  {
    "RP" : "yukifumi.kobayashi",
    "GEUK" : "han il.lee"
  },
  {
    "RP" : "ajmal.karuthedath",
    "GEUK" : "system"
  },
  {
    "RP" : "alan.owan",
    "GEUK" : "system"
  },
  {
    "RP" : "arun.sreenivas1",
    "GEUK" : "youngdug.kim"
  },
  {
    "RP" : "atsuko.matsunaga",
    "GEUK" : "system"
  },
  {
    "RP" : "bo.li",
    "GEUK" : "system"
  },
  {
    "RP" : "brian.lause",
    "GEUK" : "system"
  },
  {
    "RP" : "d.kim",
    "GEUK" : "system"
  },
  {
    "RP" : "divilin.thiyagaraj",
    "GEUK" : "youngdug.kim"
  },
  {
    "RP" : "dongwoo.lee",
    "GEUK" : "dongwoo.lee"
  },
  {
    "RP" : "douglas.stone",
    "GEUK" : "system"
  },
  {
    "RP" : "heng.zhao",
    "GEUK" : "system"
  },
  {
    "RP" : "hidekazu.masuda",
    "GEUK" : "system"
  },
  {
    "RP" : "hidekazu.masuda",
    "GEUK" : "system"
  },
  {
    "RP" : "hiroshi.hashimoto",
    "GEUK" : "system"
  },
  {
    "RP" : "jaehyeok.choi",
    "GEUK" : "jaehyeok.choi"
  },
  {
    "RP" : "justin.lanning",
    "GEUK" : "system"
  },
  {
    "RP" : "kazuyuki.dei",
    "GEUK" : "system"
  },
  {
    "RP" : "koji.miyama",
    "GEUK" : "system"
  },
  {
    "RP" : "kurt.sandstrom",
    "GEUK" : "system"
  },
  {
    "RP" : "liza.anna antony",
    "GEUK" : "TBD"
  },
  {
    "RP" : "manabu.arima",
    "GEUK" : "doo je.sung"
  },
  {
    "RP" : "manabu.arima",
    "GEUK" : "doo je.sung"
  },
  {
    "RP" : "masashi.seki",
    "GEUK" : "doo je.sung"
  },
  {
    "RP" : "matthew.bayer",
    "GEUK" : "system"
  },
  {
    "RP" : "michael.wang",
    "GEUK" : "system"
  },
  {
    "RP" : "qian.adams",
    "GEUK" : "system"
  },
  {
    "RP" : "reshma.r",
    "GEUK" : "system"
  },
  {
    "RP" : "rimon.tadross",
    "GEUK" : "system"
  },
  {
    "RP" : "ryota.torii",
    "GEUK" : "system"
  },
  {
    "RP" : "sachiyo.noguchi",
    "GEUK" : "han il.lee"
  },
  {
    "RP" : "anggeon.oh",
    "GEUK" : "anggeon.oh"
  },
  {
    "RP" : "sarath.sathikumar",
    "GEUK" : "doo je.sung"
  },
  {
    "RP" : "satoru.takahashi",
    "GEUK" : "Yujin.na"
  },
  {
    "RP" : "shunichiro.tanigawa",
    "GEUK" : "system"
  },
  {
    "RP" : "sobhita.e",
    "GEUK" : "dongyoung.choi"
  },
  {
    "RP" : "taeyun.kim",
    "GEUK" : "taeyun.kim"
  },
  {
    "RP" : "takuma.oguri",
    "GEUK" : "system"
  },
  {
    "RP" : "tetsuo.koide",
    "GEUK" : "system"
  },
  {
    "RP" : "will.youman",
    "GEUK" : "system"
  },
  {
    "RP" : "yinbo.becker",
    "GEUK" : "system"
  },
  {
    "RP" : "yuko.kanayama",
    "GEUK" : "system"
  },
  {
    "RP" : "takao.kudo",
    "GEUK" : "HW"
  },
  {
    "RP" : "tsutomu.kawaguchi1",
    "GEUK" : "HW"
  },
  {
    "RP" : "arihiro.matsumoto",
    "GEUK" : "HW"
  },
  {
    "RP" : "christin.mathew",
    "GEUK" : "taeyang.an"
  },
  {
    "RP" : "igor.stevic",
    "GEUK" : "HW"
  },
  {
    "RP" : "johny.joseph",
    "GEUK" : "V&V"
  },
  {
    "RP" : "kazuaki.okamoto",
    "GEUK" : "doo je.sung"
  },
]

ws_new = wb_new[wb_new.sheetnames[0]]  #latest weeks SPR
ws_previous = wb_old[wb_old.sheetnames[0]] #previous weeks SPR

print(ws_new)
print(ws_previous)

def search_matched_cell (value):
  for row in ws_previous.iter_rows(min_row=1, min_col=2, max_col=2):
      for cell in row:
        if(ws_previous[cell.coordinate].value == value): 
          return cell.row  

  return 99999

def mark_newly_closed_cell (newcellrow, cellrow, new_value):
  for row in ws_previous.iter_rows(min_row=cellrow, max_row=cellrow, min_col=13, max_col=13):  # status column
    for cell in row:
        pre_val = ws_previous["M"+str(cellrow)].value
        if (pre_val != new_value):
          if((new_value == "Verified") or (new_value == "Resolved") or (new_value == "Closed")):
            ws_new["M"+str(newcellrow)].fill = PatternFill(fgColor="00FF00", fill_type = "solid")


for row in ws_previous.iter_rows(min_row=1,min_col=2, max_col=2):
  for cell in row:
    cell_row = search_matched_cell (ws_new["B"+str(cell.row)].value)
    if(cell_row == 99999):
      print("cell not found for", ws_new["B"+str(cell.row)].value )
    else:
      mark_newly_closed_cell (cell.row, cell_row, ws_new["M"+str(cell.row)].value)   # Status column
      ws_new["AB"+str(cell.row)].value = ws_previous["AB"+str(cell_row)].value
      ws_new["AB"+str(cell.row)].font = copy(ws_previous["AB"+str(cell_row)].font)
      ws_new["AB"+str(cell.row)].border = copy(ws_previous["AB"+str(cell_row)].border)
      ws_new["AB"+str(cell.row)].fill = copy(ws_previous["AB"+str(cell_row)].fill)
      ws_new["AB"+str(cell.row)].number_format = copy(ws_previous["AB"+str(cell_row)].number_format)
      ws_new["AB"+str(cell.row)].protection = copy(ws_previous["AB"+str(cell_row)].protection)
      ws_new["AB"+str(cell.row)].alignment = copy(ws_previous["AB"+str(cell_row)].alignment)
        
      ws_new["AC"+str(cell.row)].value = ws_previous["AC"+str(cell_row)].value
      ws_new["AC"+str(cell.row)].font = copy(ws_previous["AC"+str(cell_row)].font)
      ws_new["AC"+str(cell.row)].border = copy(ws_previous["AC"+str(cell_row)].border)
      ws_new["AC"+str(cell.row)].fill = copy(ws_previous["AC"+str(cell_row)].fill)
      ws_new["AC"+str(cell.row)].number_format = copy(ws_previous["AC"+str(cell_row)].number_format)
      ws_new["AC"+str(cell.row)].protection = copy(ws_previous["AC"+str(cell_row)].protection)
      ws_new["AC"+str(cell.row)].alignment = copy(ws_previous["AC"+str(cell_row)].alignment)

      ws_new["AD"+str(cell.row)].value = ws_previous["AD"+str(cell_row)].value
      ws_new["AD"+str(cell.row)].font = copy(ws_previous["AD"+str(cell_row)].font)
      ws_new["AD"+str(cell.row)].border = copy(ws_previous["AD"+str(cell_row)].border)
      ws_new["AD"+str(cell.row)].fill = copy(ws_previous["AD"+str(cell_row)].fill)
      ws_new["AD"+str(cell.row)].number_format = copy(ws_previous["AD"+str(cell_row)].number_format)
      ws_new["AD"+str(cell.row)].protection = copy(ws_previous["AD"+str(cell_row)].protection)
      ws_new["AD"+str(cell.row)].alignment = copy(ws_previous["AD"+str(cell_row)].alignment)

      ws_new["AE"+str(cell.row)].value = ws_previous["AE"+str(cell_row)].value
      ws_new["AE"+str(cell.row)].font = copy(ws_previous["AE"+str(cell_row)].font)
      ws_new["AE"+str(cell.row)].border = copy(ws_previous["AE"+str(cell_row)].border)
      ws_new["AE"+str(cell.row)].fill = copy(ws_previous["AE"+str(cell_row)].fill)
      ws_new["AE"+str(cell.row)].number_format = copy(ws_previous["AE"+str(cell_row)].number_format)
      ws_new["AE"+str(cell.row)].protection = copy(ws_previous["AE"+str(cell_row)].protection)
      ws_new["AE"+str(cell.row)].alignment = copy(ws_previous["AE"+str(cell_row)].alignment)

# Reviewer in GEUK	Status
ws_new.column_dimensions[get_column_letter(28)].width = ws_previous.column_dimensions[get_column_letter(27)].width    
ws_new.column_dimensions[get_column_letter(29)].width = ws_previous.column_dimensions[get_column_letter(28)].width    


for row in ws_new.iter_rows(min_row=1, min_col=28, max_col=29):
    for cell in row:
        if(((ws_new[cell.coordinate].value == None) or (ws_new[cell.coordinate].value == 'TBD')) and cell.column == 28):
            for name in total_name_table:
                if (name["RP"] == ws_new["N"+str(cell.row)].value):
                    ws_new[cell.coordinate].value = name["GEUK"]
                    break

wb_new.save(new_file_name)
wb_old.save(old_file_name)

wb_new.close()
wb_old.close()

