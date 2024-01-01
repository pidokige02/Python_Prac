from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill

from copy import copy, deepcopy
import sys

new_file_name = "osprey_issues_new.xlsx"
old_file_name = "osprey_issues_old.xlsx"

if len(sys.argv) == 1:
    txt = "running : Python xl_smart.py {} {}".format(old_file_name, new_file_name)
    print(txt) 
elif len(sys.argv) == 2:
    print("Usage : Python xl_smart.py oldfilename newfilename!")
    txt = "running : Python xl_smart.py {} {}".format(old_file_name, new_file_name)
    print(txt) 
    sys.exit()
elif len(sys.argv) == 3:
    old_file_name = sys.argv[1]
    new_file_name = sys.argv[2]
    txt = "running : Python xl_smart.py {} {}".format(old_file_name, new_file_name)
    print(txt) 
else:  
    print("Usage : Python xl_smart.py oldfilename newfilename!")
    sys.exit()

wb_old = load_workbook(old_file_name)
wb_new = load_workbook(new_file_name)

print(old_file_name)
print(new_file_name)

print(wb_new.sheetnames)
print(wb_old.sheetnames)


ws_new = wb_new[wb_new.sheetnames[0]]  #latest weeks SPR
ws_previous = wb_old[wb_old.sheetnames[0]] #previous weeks SPR

print(ws_new)
print(ws_previous)

def search_matched_cell (value):
  for row in ws_previous.iter_rows(min_row=1, min_col=2, max_col=2):
      for cell in row:
        if(ws_previous[cell.coordinate].value == value): 
          return cell.row  

  return 99999

def mark_newly_closed_cell (newcellrow, cellrow, new_value):
  for row in ws_previous.iter_rows(min_row=cellrow, max_row=cellrow, min_col=13, max_col=13):  # status column
    for cell in row:
        pre_val = ws_previous["M"+str(cellrow)].value
        if (pre_val != new_value):
          if((new_value == "Verified") or (new_value == "Resolved") or (new_value == "Closed")):
            ws_new["M"+str(newcellrow)].fill = PatternFill(fgColor="00FF00", fill_type = "solid")

# def mark_newly_closed_cell (cellrow, new_value):
#   for row in ws_previous.iter_rows(min_row=cellrow, max_row=cellrow, min_col=13, max_col=13):  # status column
#     for cell in row:
#         pre_val = ws_previous["M"+str(cellrow)].value
#         if (pre_val != new_value):
#           if((new_value == "Verified") or (new_value == "Resolved") or (new_value == "Closed")):
#             ws_new["M"+str(cellrow)].fill = PatternFill(fgColor="00FF00", fill_type = "solid")

for row in ws_previous.iter_rows(min_row=1,min_col=2, max_col=2):
  for cell in row:
    cell_row = search_matched_cell (ws_new["B"+str(cell.row)].value)
    if(cell_row == 99999):
      print("cell not found for", ws_new["B"+str(cell.row)].value )
    else:
      mark_newly_closed_cell (cell.row, cell_row, ws_new["M"+str(cell.row)].value)   # Status column
      ws_new["AB"+str(cell.row)].value = ws_previous["AB"+str(cell_row)].value
      ws_new["AB"+str(cell.row)].font = copy(ws_previous["AB"+str(cell_row)].font)
      ws_new["AB"+str(cell.row)].border = copy(ws_previous["AB"+str(cell_row)].border)
      ws_new["AB"+str(cell.row)].fill = copy(ws_previous["AB"+str(cell_row)].fill)
      ws_new["AB"+str(cell.row)].number_format = copy(ws_previous["AB"+str(cell_row)].number_format)
      ws_new["AB"+str(cell.row)].protection = copy(ws_previous["AB"+str(cell_row)].protection)
      ws_new["AB"+str(cell.row)].alignment = copy(ws_previous["AB"+str(cell_row)].alignment)
        
      ws_new["AC"+str(cell.row)].value = ws_previous["AC"+str(cell_row)].value
      ws_new["AC"+str(cell.row)].font = copy(ws_previous["AC"+str(cell_row)].font)
      ws_new["AC"+str(cell.row)].border = copy(ws_previous["AC"+str(cell_row)].border)
      ws_new["AC"+str(cell.row)].fill = copy(ws_previous["AC"+str(cell_row)].fill)
      ws_new["AC"+str(cell.row)].number_format = copy(ws_previous["AC"+str(cell_row)].number_format)
      ws_new["AC"+str(cell.row)].protection = copy(ws_previous["AC"+str(cell_row)].protection)
      ws_new["AC"+str(cell.row)].alignment = copy(ws_previous["AC"+str(cell_row)].alignment)

      ws_new["AD"+str(cell.row)].value = ws_previous["AD"+str(cell_row)].value
      ws_new["AD"+str(cell.row)].font = copy(ws_previous["AD"+str(cell_row)].font)
      ws_new["AD"+str(cell.row)].border = copy(ws_previous["AD"+str(cell_row)].border)
      ws_new["AD"+str(cell.row)].fill = copy(ws_previous["AD"+str(cell_row)].fill)
      ws_new["AD"+str(cell.row)].number_format = copy(ws_previous["AD"+str(cell_row)].number_format)
      ws_new["AD"+str(cell.row)].protection = copy(ws_previous["AD"+str(cell_row)].protection)
      ws_new["AD"+str(cell.row)].alignment = copy(ws_previous["AD"+str(cell_row)].alignment)

# for row in ws_previous.iter_rows(min_row=1, min_col=28, max_col=30):
#     for cell in row:
#         cell_row = search_matched_cell (ws_new["B"+str(cell.row)].value)
#         if(cell_row == 99999):
#             print("cell not found for", ws_new["B"+str(cell.row)].value )
#         else:
#             ws_new[cell.coordinate].value = ws_previous[cell.coordinate].value
#             ws_new[cell.coordinate].font = copy(ws_previous[cell.coordinate].font)
#             ws_new[cell.coordinate].border = copy(ws_previous[cell.coordinate].border)
#             ws_new[cell.coordinate].fill = copy(ws_previous[cell.coordinate].fill)
#             ws_new[cell.coordinate].number_format = copy(ws_previous[cell.coordinate].number_format)
#             ws_new[cell.coordinate].protection = copy(ws_previous[cell.coordinate].protection)
#             ws_new[cell.coordinate].alignment = copy(ws_previous[cell.coordinate].alignment)
    
#             mark_newly_closed_cell(cell.row, ws_new["M"+str(cell.row)].value)

# Critical	Triage	Comments column
ws_new.column_dimensions[get_column_letter(28)].width = ws_previous.column_dimensions[get_column_letter(27)].width    
ws_new.column_dimensions[get_column_letter(29)].width = ws_previous.column_dimensions[get_column_letter(28)].width    
ws_new.column_dimensions[get_column_letter(30)].width = ws_previous.column_dimensions[get_column_letter(29)].width    

# change sheet name of new comsolidated file
position = new_file_name.find('FW')
# print ('position', position)
# print ('slice', new_file_name[position:position+4])

ws_new.title = new_file_name[position:position+4] # slice FWXX and change sheet title w/ it

wb_new.save(new_file_name)
wb_old.save(old_file_name)
wb_new.close()
wb_old.close()

