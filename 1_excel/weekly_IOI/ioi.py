from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.page import PageMargins
import pandas as pd 

from copy import copy, deepcopy
import sys, os
from datetime import datetime


exp_file_name = 'export.csv'
excel_file_name = 'export.xlsx'
new_excel_file_name = 'copied.xlsx'

column_dimension_table = {
    'A' : 7.0, 
    'B' : 8.0,
    'C' : 8.0,
    'D' : 21.0,
    'E' : 40.0,
    'F' : 42.0,
    'G' : 7.0,
    'H' : 8.0,
    'I' : 8.0
}

column_name_table = {
    'A1' : "Flag", 
    'B1' : "Tags",
    'C1' : "ID",
    'D1' : "Name",
    'E1' : "Desc",
    'F1' : "Notes",
    'G1' : "Status",
    'H1' : "Owner",
    'I1' : "Due"
}

product_search_table = {
    "Osprey R4" : 1,
    "Osprey R5" : 2,
    "Osprey2mainline" : 3,
    "Gemini R4" : 4,
    "Gemini R5" : 5,
    "Maru R4.5" : 6,
    "Maru R4" : 7,
    "Maru R3" : 8,
    "Others" : 9,
    "Advantech" : 10
}


def set_page_properties (sheet):
    # Set the page orientation to landscape
    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    #  Set the page size to PAPERSIZE_A4
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4

    #narrow_margin 설정
    sheet.page_margins = PageMargins (left=0.25, right=0.25, top=0.75, bottom=0.75)


def set_page_area (sheet, row_count):
    # # flexible setting of print area based on real captured data.
    print_area = "B1:I{}".format(row_count)
    sheet.print_area = print_area


def set_cells_properties(sheet, minrow, mincol, maxcol):

    bold_font = Font(bold=True)
    #set column demension with predefined one.
    for column in sheet.iter_cols(min_row=1, max_row=1, min_col=1, max_col=maxcol):
        column_letter = column[0].column_letter
        sheet.column_dimensions[column_letter].width = column_dimension_table[column_letter]
        for cell in column:     # make title bold
            cell.font = bold_font

    
    for row in sheet.iter_rows(min_row=minrow, min_col=mincol, max_col=maxcol):
        for cell in row:
            # Center the content both horizontally and vertically
            # long text will wrap to the next line
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)


def csv2excel(filename):
    csv_file_path = filename
    # need to chek file ext type

    print(f"Running: python ioi.py {csv_file_path}")

    # CSV 파일 읽기
    df = pd.read_csv(csv_file_path)

    report_file_name = excel_file_name

    if os.path.exists(report_file_name):
        print(f"{report_file_name} 파일이 존재합니다.")
        try:
            os.remove(report_file_name)
            print(f"기존 {report_file_name} 파일이 삭제되었습니다.")
        except OSError as e:
            print(f"파일 삭제 실패: {e}")

    print(f"{report_file_name} 만들겠습니다.")
    df.to_excel(report_file_name, index=False)

    wb_report = load_workbook(report_file_name)

    return wb_report


def creat_blank_column(column_to_insert=1, number_to_insert=2):
    report_sheet.insert_cols(column_to_insert,number_to_insert)
    

def move_columns (start_column_char, end_column_char, max_row, shift_number):
    cell_range = "{0}1:{1}{2}".format(start_column_char,end_column_char, max_row)
    # print("cell_range", cell_range)
    report_sheet.move_range(cell_range, rows=0, cols=shift_number)


def rearrange_columns(sheet):
    column_to_insert = 1  # "A"
    number_to_insert = 2  # "A" and "B" 
    char_Expedite_column = "E" # if two columns are inserted original 'E' has Expedite 
    char_Tags_column = "F"  # # if two columns are inserted original 'E' has Expedite
    shift_number = -4 # shfit back -4
    delete_column = 5 # "E" 
    delete_num = 2 # "E" and "F"
    # 빈 열 추가 in order to move Tags and Expertide column, .
    creat_blank_column(column_to_insert, number_to_insert)
    move_columns (char_Expedite_column, char_Tags_column, row_count, shift_number)
    sheet.delete_cols(delete_column, delete_num)
    

def simple_datatime(date_string):
    # 입력된 날짜를 FWXX.X format 으로 변경
    date_object = datetime.strptime(date_string, "%Y-%m-%d")

    formatted_date = date_object.strftime("FW%W.%w") 

    return formatted_date


if len(sys.argv) != 2:
    print("Usage: python xl_weekly_ioi.py <filename>")
    sys.exit()

try:
    wb_report = csv2excel(exp_file_name)

    report_sheet = wb_report[wb_report.sheetnames[0]]

    row_count = report_sheet.max_row
    max_column = report_sheet.max_column

    column_to_insert = 1
    number_to_insert = 2 

    print(f"The number of rows in the sheet is: {row_count}")
    print(f"max_column: {max_column}")

    set_page_properties (report_sheet)
    rearrange_columns(report_sheet)

    # if not merged with "Delivery Date" and "Target Data", 
    # merge them with "Due"
    cell_value1 = report_sheet['I1'].value  # for Target Date
    cell_value2 = report_sheet['J1'].value  # for Delivery Date

    if ((cell_value1 != "Target Date") and (cell_value2 != "Delivery Date")):
        print("excel file has wrong format. Can't process it")
        sys.exit() 

    # combine "I" and "J" into one column
    for row in report_sheet.iter_rows(min_row=2, max_row=row_count, min_col=9, max_col=9):
        for cell in row:
            copy_val = report_sheet["J"+str(cell.row)].value
            target_val = report_sheet[cell.coordinate].value 
            if ((copy_val is not None) and (target_val is None)) :
                report_sheet[cell.coordinate].value = copy_val
    
    del_column = 10  # "J" column
    report_sheet.delete_cols(del_column, 1)

    # change due data format into simpler one
    for row in report_sheet.iter_rows(min_row=2, max_row=row_count, min_col=9, max_col=9):
        for cell in row:
            target_date = report_sheet[cell.coordinate].value
            if ((target_date is not None)):
                report_sheet[cell.coordinate].value = simple_datatime(target_date)


    # update row_count, max_column 
    row_count = report_sheet.max_row
    max_column = report_sheet.max_column

    set_page_area (report_sheet, row_count)

    # change title name to simplified one
    for row in report_sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=9):
        for cell in row:
            report_sheet[cell.coordinate].value = column_name_table[cell.coordinate] 

    # # Constants for better readability
    MIN_ROW, MIN_COL, MAX_COL = 1, 1, max_column
    set_cells_properties(report_sheet, MIN_ROW, MIN_COL, MAX_COL)

    # 새로운 excel file 생성
    new_wb = Workbook()

    # #새로운 sheet 생성
    new_sheet = new_wb.active

    # append title rowl
    for row in report_sheet.iter_rows(min_row=1, max_row=1, values_only=True):
        new_sheet.append(row)

    # sort tag column 
    for row in report_sheet.iter_rows(min_row=2, max_row=row_count, min_col=2, max_col=2):
        for cell in row:
            indexlist = []
            resultdict = {} 
            product_string = report_sheet[cell.coordinate].value
            result_list = product_string.split(";")
            for x in result_list:
                indexlist.append(product_search_table[x])   # product_search_table 와 같은 dict 를 만들기위한 1,2,3,4  를 만들기 위한 index list
                resultdict = dict(zip(result_list, indexlist))  # product_search_table 와 같은 dict 를 만드는 과정임
                length_of_dict = len(resultdict)
                if(length_of_dict > 1):  # sorting 이 필요한지 확인하는 것임
                    sorted_dict_by_value = dict(sorted(resultdict.items(), key=lambda item: item[1]))  # sorting 된 것임
                    keys_list = list(sorted_dict_by_value.keys())  # key 값 osprey R4, osprey R5 와 같은 list 를 만드는 것암
                    result_string = ';'.join(keys_list)  # ;으로 분리된  string (ex : Osprey R4; Osprey2mainline)
                    report_sheet[cell.coordinate].value = result_string


    for prouct in product_search_table:
        for row in report_sheet.iter_rows(min_row=1, max_row=row_count, min_col=MIN_COL, max_col=MAX_COL, values_only=True):
            for col_num, value in enumerate(row, start=1):
                cell_address = f"{report_sheet.cell(row=report_sheet.min_row, column=col_num).column_letter}{report_sheet.cell(row=report_sheet.min_row, column=col_num).row}"
                if(cell_address == 'B1'):  # tag column that contain product data (Osprey R4, Osprey R5 ex)
                    product_string = str(value)
                    result_list = product_string.split(";")
                    # print(result_list)
                    # print(f"셀 위치: {cell_address}, 값: {value}")
                    for index, value in enumerate(result_list):
                        if (index  == 0 and value == prouct):
                            new_sheet.append(row)


    set_page_properties (new_sheet)
    set_page_area (new_sheet, row_count)
    set_cells_properties(new_sheet, MIN_ROW, MIN_COL, MAX_COL)
           
    # # Save the modified workbook
    wb_report.save(excel_file_name)
    new_wb.save(new_excel_file_name)


finally:
    # Ensure the workbook is closed, even if an exception occurs
    if 'wb_report' in locals():
        wb_report.close()
    if 'new_wb' in locals():
        new_wb.close()

