import smartsheet
from smartsheet import Smartsheet
from smartsheet import sheets

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.page import PageMargins, PrintOptions
from bs4 import BeautifulSoup 
from openpyxl.utils import get_column_letter

from copy import copy, deepcopy
import sys, os
from datetime import datetime
from .Utils import *


excel_file_name = 'Data/export.xlsx'
template_excel_file_name = 'Data/template.xlsx'
sheet_excel_file_name = 'Data/sheetexport.xlsx'

# 들어오는 SPR DATA 를 excel file 로 변경하는 험수 
def spr2excel(sprs, arg_alm_table_header, arg_alm_table_map):

    report_file_name = excel_file_name

    if os.path.exists(report_file_name):
        print(f"{report_file_name} 파일이 존재합니다.")
        try:
            os.remove(report_file_name)
            print(f"기존 {report_file_name} 파일이 삭제되었습니다.")
        except OSError as e:
            print(f"파일 삭제 실패: {e}")
    
    try:
        column_width_table = {}

        template_wb = load_workbook(template_excel_file_name)
        new_wb = Workbook()

        # #새로운 sheet 생성
        new_sheet = new_wb.active
        template_sheet = template_wb.active

        # get maxcolumn 
        maxcolumn = len(arg_alm_table_header) 
        additional_columns = 3

        # append title rowl
        for col, value in arg_alm_table_header.items():
            new_sheet[col+'1'] = value[0]

        # add data into below row    
        for spr in sprs:
            headers = list(spr.keys())
            sorted_columns = sorted(headers, key=lambda x: list(arg_alm_table_map.keys()).index(x))  # sort it for exported excel
            row_data = [spr[header] for header in sorted_columns]
            new_sheet.append(row_data)

        # get dimension based on template excel file
        for column in template_sheet.iter_cols(min_row=1, max_row=1, min_col=1, max_col=maxcolumn):
            column_letter = column[0].column_letter
            column_width_table[column_letter] = template_sheet.column_dimensions[column_letter].width 

        # set dimension based on template excel file
        for column in new_sheet.iter_cols(min_row=1, max_row=1, min_col=1, max_col=maxcolumn):
            column_letter = column[0].column_letter
            new_sheet.column_dimensions[column_letter].width = column_width_table[column_letter]

        # set html element removed content for DRB session
        for row in new_sheet.iter_rows(min_row=2,min_col=maxcolumn):
            for cell in row:
                html_content = new_sheet[cell.coordinate].value
                soup = BeautifulSoup(html_content, 'html.parser')
                visible_text = soup.get_text(separator=' ', strip=True)
                new_sheet[cell.coordinate].value = visible_text

        # set the property of title,  3 is additional three column
        for row in new_sheet.iter_rows(min_row=1,max_row=1, min_col=1, max_col=maxcolumn+additional_columns):
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)
                new_sheet[cell.coordinate].fill = copy(template_sheet[cell.coordinate].fill)
                new_sheet[cell.coordinate].font = copy(template_sheet[cell.coordinate].font)
                new_sheet[cell.coordinate].border = copy(template_sheet[cell.coordinate].border)

        # # set the property of spr contents, 3 is additional three column
        for row in new_sheet.iter_rows(min_row=2,min_col=1, max_col=maxcolumn+additional_columns):
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)
                new_sheet[cell.coordinate].font = copy(template_sheet[get_column_letter(cell.column)+"2"].font)
                new_sheet[cell.coordinate].border = copy(template_sheet[get_column_letter(cell.column)+"2"].border)
                if(cell.row % 2 == 0 ):
                    new_sheet[cell.coordinate].fill = copy(template_sheet[get_column_letter(cell.column)+"2"].fill)
                else:
                    new_sheet[cell.coordinate].fill = copy(template_sheet[get_column_letter(cell.column)+"3"].fill)

        # # set the property of number format to text for column 'B' to remove warning
        for row in new_sheet.iter_rows(min_row=2,min_col=2, max_col=2):
            for cell in row:
                new_sheet[cell.coordinate].number_format = '@'

        # 열(column) 숨기기
        for column, data in arg_alm_table_header.items():
            new_sheet.column_dimensions[column].hidden = data[2]  # data[2] has hide flag

        # # Save the modified workbook
        new_wb.save(excel_file_name)
        template_wb.save(template_excel_file_name)

    except FileNotFoundError:
        print("파일을 찾을 수 없습니다.")
    except openpyxl.utils.exceptions.InvalidFileException:
        print("올바른 형식의 엑셀 파일이 아닙니다.")
    except Exception as e:
        print("알 수 없는 오류가 발생했습니다:", e)

    finally:
        # Ensure the workbook is closed, even if an exception occurs
        if 'new_wb' in locals():
            new_wb.close()
        if 'template_wb' in locals():
            template_wb.close()


#smartsheet 에 있는 SPR data 를 excel 로 만들어 내는 함수.
def sheetspr2excel(sprs, smart_table_header):

    report_file_name = sheet_excel_file_name

    if os.path.exists(report_file_name):
        print(f"{report_file_name} 파일이 존재합니다.")
        try:
            os.remove(report_file_name)
            print(f"기존 {report_file_name} 파일이 삭제되었습니다.")
        except OSError as e:
            print(f"파일 삭제 실패: {e}")
    
    try:
        column_width_table = {}

        template_wb = load_workbook(template_excel_file_name)
        new_wb = Workbook()

        # #새로운 sheet 생성
        new_sheet = new_wb.active
        template_sheet = template_wb.active

        # get maxcolumn 
        maxcolumn = len(smart_table_header) 

        # append title rowl
        for col, value in smart_table_header.items():
            new_sheet[col+'1'] = value[0]

        # add data ito below row    
        for spr in sprs:
            new_sheet.append(spr)

        # # Save the modified workbook
        new_wb.save(report_file_name)
        template_wb.save(template_excel_file_name)

    except FileNotFoundError:
        print("파일을 찾을 수 없습니다.")
    except openpyxl.utils.exceptions.InvalidFileException:
        print("올바른 형식의 엑셀 파일이 아닙니다.")
    except Exception as e:
        print("알 수 없는 오류가 발생했습니다:", e)

    finally:
        # Ensure the workbook is closed, even if an exception occurs
        if 'new_wb' in locals():
            new_wb.close()
        if 'template_wb' in locals():
            template_wb.close()

 
# excel 로 있는  SPR 을 smartsheet 로 올리는 함수
def excel2sheet(smartsheet_client, configData, report_file_name, product):

    print("Start uploading...")

    absolute_path = get_absolate_path(report_file_name)
    filename = os.path.basename(absolute_path)
    file_name_without_extension = os.path.splitext(filename)[0]

    print("abspath :" + absolute_path)
    print("filename : " + filename)
    print("file_name_without_extension : " + file_name_without_extension)


    # 워크스페이스 안의 시트 목록 가져오기
    workspaces = smartsheet_client.Workspaces.list_workspaces(include_all=True)

    # 작업 공간 및 ID 출력
    for workspace in workspaces.data:
        if(workspace.name == "GEUK SW Eng."):
            # Get the workspace
            response = smartsheet_client.Workspaces.get_workspace(workspace.id)
            print("Workspace Name:", workspace.name)
            # print("Workspace ID:", workspace.id)

            # Folder 목록 출력
            for folder in response.to_dict()['folders']:
                if(folder['name'] == "SPR Tracking"):
                    print("Folder Name:", folder['name'])
                    # print("Folder ID:", folder['id'])

                    response2 = smartsheet_client.Folders.get_folder(folder['id'])

                    # 이미 같은 이름의 sheet 가 있으면 지운다.
                    for sheet in response2.to_dict()['sheets']:
                        if(sheet['name'] == file_name_without_extension):
                            print("Deleting existing Sheet Name:", sheet['name'])
                            # print("Sheet ID:", sheet['id'])
                            smartsheet_client.Sheets.delete_sheet(sheet['id'])

                    # Import the sheet
                    result = smartsheet_client.Folders.import_xlsx_sheet(folder['id'], absolute_path, header_row_index=0)

                    # Load entire sheet
                    sheet = smartsheet_client.Sheets.get_sheet(result.data.id)
                    # print("smartsheetID : " + str(sheet.id))
                    print("Loaded " + str(len(sheet.rows)) + " rows from sheet: " + sheet.name)

                    if(product == "Osprey R4"):
                        configData.update({"SmartsheetID1": str(sheet.id)})  # newly save sheetID1
                    elif(product == "Gemini R5"):
                        configData.update({"SmartsheetID2": str(sheet.id)})  # newly save sheetID2
                    elif(product == "Gemini R4"):
                        configData.update({"SmartsheetID3": str(sheet.id)})  # newly save sheetID3
                    elif(product == "Gemini R3"):
                        configData.update({"SmartsheetID4": str(sheet.id)})  # newly save sheetID4
                    else:
                        configData.update({"SmartsheetID1": str(sheet.id)})  # newly save sheetID1

                    writeConfigData(configData)

                    print("Done")

    return  sheet.id


def create_product_file(product):

    date = get_current_date()
    fw_format = date_to_fw_format (date)

    if(product == "Osprey R4"):
        new_file_name = f"Data/osprey_issues_2024_{fw_format}.xlsx"
    elif(product == "Gemini R5"):
        new_file_name = f"Data/Gemini_R5_issues_2024_{fw_format}.xlsx"
    elif(product == "Gemini R4"):
        new_file_name = f"Data/Gemini_R4_issues_2024_{fw_format}.xlsx"
    elif(product == "Osprey R4 clone"):
        new_file_name = f"Data/osprey_link_2024_{fw_format}.xlsx"
    elif(product == "Gemini R3"):
        new_file_name = f"Data/Gemini_R3_issues_2024_{fw_format}.xlsx"
    else:
        new_file_name = f"Data/osprey_issues_2024_{fw_format}.xlsx"

    file_name_change(excel_file_name, new_file_name)

    return new_file_name