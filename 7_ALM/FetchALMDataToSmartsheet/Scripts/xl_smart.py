from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill

import re

from copy import copy, deepcopy
import sys
from datetime import datetime
# subprocess.run(["python", "./Scripts/xl_smart.py", "./Data/sheetexport.xlsx", "./Data/export.xlsx"])
# 을 실행하였을때 ./Scripts/xl_smart.py 안의 from .Utils import * 에서  ImportError: attempted relative import with no known parent package 가 발생합니다.# 

def get_current_date():
    # 현재 날짜와 시간 얻기
    current_date_time = datetime.now()    
    
    # 현재 날짜만 얻기
    current_date = current_date_time.date()

    return current_date

def get_current_year():
    # 현재 날짜와 시간 얻기
    current_date_time = datetime.now()    
    
    # 현재 날짜만 얻기
    current_year = str(current_date_time.year)

    return current_year

def date_to_fw_format(input_date):
    # 입력된 날짜를 파싱
    date_object = datetime.strptime(str(input_date), '%Y-%m-%d')

    # 날짜에서 년도와 주차(WW) 추출
    year = date_object.year
    week_number = date_object.strftime('%U')
    week_number = int(week_number)+1 # start FW01 not FW00 
    str_week_number = str(week_number)

    # FWxx 형식으로 포맷팅
    fw_format = f'FW{str_week_number.zfill(2)}'

    return fw_format

def analysis_drb(text, pattern):
   
   matches = re.findall(pattern, text, re.IGNORECASE)
   
   return matches

   
def put_issue_triage():  # based on DRB, isseu will be categorized as minor major, IO
  for row in ws_new.iter_rows(min_row=2,min_col=27, max_col=27):  # 27 은 DRB column 을 의미함. 
    for cell in row:
      if ws_new["J"+str(cell.row)].value == "NC - Design Non-Conformance" and  ws_new["AD"+str(cell.row)].value == "IO":   # wrong setting
        ws_new["AD"+str(cell.row)].value = None

      if ws_new["J"+str(cell.row)].value == "IO - Improvement Opportunity" and  ws_new["AD"+str(cell.row)].value is None:
        ws_new["AD"+str(cell.row)].value = "IO"
      elif ws_new["J"+str(cell.row)].value == "NC - Design Non-Conformance" and  ws_new["AD"+str(cell.row)].value is None:
        if (ws_new["AA"+str(cell.row)].value):
          match = analysis_drb(str(ws_new["AA"+str(cell.row)].value), r"#Minor#")
          if(match):
              ws_new["AD"+str(cell.row)].value = "Minor"
          match = analysis_drb(str(ws_new["AA"+str(cell.row)].value), r"#Major#")
          if(match):
              ws_new["AD"+str(cell.row)].value = "Major"         
      else:
        pass           


new_file_name = "osprey_issues_new.xlsx"
old_file_name = "osprey_issues_old.xlsx"

if len(sys.argv) == 1:
    txt = "running : Python xl_smart.py {} {}".format(old_file_name, new_file_name)
    print(txt) 
elif len(sys.argv) == 2:
    print("Usage : Python xl_smart.py oldfilename newfilename!")
    txt = "running : Python xl_smart.py {} {}".format(old_file_name, new_file_name)
    print(txt) 
    sys.exit()
elif len(sys.argv) == 3:
    old_file_name = sys.argv[1]
    new_file_name = sys.argv[2]
    txt = "running : Python xl_smart.py {} {}".format(old_file_name, new_file_name)
    print(txt) 
else:  
    print("Usage : Python xl_smart.py oldfilename newfilename!")
    sys.exit()


wb_old = load_workbook(old_file_name)
wb_new = load_workbook(new_file_name)

print(old_file_name)
print(new_file_name)

print(wb_new.sheetnames)
print(wb_old.sheetnames)


ws_new = wb_new[wb_new.sheetnames[0]]  #latest weeks SPR
ws_previous = wb_old[wb_old.sheetnames[0]] #previous weeks SPR

print(ws_new)
print(ws_previous)

def search_matched_cell (value):
  for row in ws_previous.iter_rows(min_row=2, min_col=2, max_col=2):
      for cell in row:
        strval = str(ws_previous[cell.coordinate].value)  # jinha correct FW05.1
        if(strval == value): 
          return cell.row  

  return 99999

def mark_newly_closed_cell (newcellrow, cellrow, new_value):
  for row in ws_previous.iter_rows(min_row=cellrow, max_row=cellrow, min_col=13, max_col=13):  # status column
    for cell in row:
        pre_val = ws_previous["M"+str(cellrow)].value
        if (pre_val != new_value):
          if((new_value == "Verified") or (new_value == "Resolved") or (new_value == "Closed")):
            ws_new["M"+str(newcellrow)].fill = PatternFill(fgColor="00FF00", fill_type = "solid")

# def mark_newly_closed_cell (cellrow, new_value):
#   for row in ws_previous.iter_rows(min_row=cellrow, max_row=cellrow, min_col=13, max_col=13):  # status column
#     for cell in row:
#         pre_val = ws_previous["M"+str(cellrow)].value
#         if (pre_val != new_value):
#           if((new_value == "Verified") or (new_value == "Resolved") or (new_value == "Closed")):
#             ws_new["M"+str(cellrow)].fill = PatternFill(fgColor="00FF00", fill_type = "solid")

# copy header dada
ws_new["AB1"].value = ws_previous["AB1"].value
ws_new["AC1"].value = ws_previous["AC1"].value
ws_new["AD1"].value = ws_previous["AD1"].value
             
for row in ws_previous.iter_rows(min_row=2,min_col=2, max_col=2):
  for cell in row:
    cell_row = search_matched_cell (ws_new["B"+str(cell.row)].value)
    if(cell_row == 99999):
      print("cell not found for", ws_new["B"+str(cell.row)].value )
    else:
      mark_newly_closed_cell (cell.row, cell_row, ws_new["M"+str(cell.row)].value)   # Status column
      ws_new["AB"+str(cell.row)].value = ws_previous["AB"+str(cell_row)].value
      ws_new["AB"+str(cell.row)].font = copy(ws_previous["AB"+str(cell_row)].font)
      ws_new["AB"+str(cell.row)].border = copy(ws_previous["AB"+str(cell_row)].border)
      ws_new["AB"+str(cell.row)].fill = copy(ws_previous["AB"+str(cell_row)].fill)
      ws_new["AB"+str(cell.row)].number_format = copy(ws_previous["AB"+str(cell_row)].number_format)
      ws_new["AB"+str(cell.row)].protection = copy(ws_previous["AB"+str(cell_row)].protection)
      ws_new["AB"+str(cell.row)].alignment = copy(ws_previous["AB"+str(cell_row)].alignment)
        
      ws_new["AC"+str(cell.row)].value = ws_previous["AC"+str(cell_row)].value
      ws_new["AC"+str(cell.row)].font = copy(ws_previous["AC"+str(cell_row)].font)
      ws_new["AC"+str(cell.row)].border = copy(ws_previous["AC"+str(cell_row)].border)
      ws_new["AC"+str(cell.row)].fill = copy(ws_previous["AC"+str(cell_row)].fill)
      ws_new["AC"+str(cell.row)].number_format = copy(ws_previous["AC"+str(cell_row)].number_format)
      ws_new["AC"+str(cell.row)].protection = copy(ws_previous["AC"+str(cell_row)].protection)
      ws_new["AC"+str(cell.row)].alignment = copy(ws_previous["AC"+str(cell_row)].alignment)

      ws_new["AD"+str(cell.row)].value = ws_previous["AD"+str(cell_row)].value
      ws_new["AD"+str(cell.row)].font = copy(ws_previous["AD"+str(cell_row)].font)
      ws_new["AD"+str(cell.row)].border = copy(ws_previous["AD"+str(cell_row)].border)
      ws_new["AD"+str(cell.row)].fill = copy(ws_previous["AD"+str(cell_row)].fill)
      ws_new["AD"+str(cell.row)].number_format = copy(ws_previous["AD"+str(cell_row)].number_format)
      ws_new["AD"+str(cell.row)].protection = copy(ws_previous["AD"+str(cell_row)].protection)
      ws_new["AD"+str(cell.row)].alignment = copy(ws_previous["AD"+str(cell_row)].alignment)

# for row in ws_previous.iter_rows(min_row=1, min_col=28, max_col=30):
#     for cell in row:
#         cell_row = search_matched_cell (ws_new["B"+str(cell.row)].value)
#         if(cell_row == 99999):
#             print("cell not found for", ws_new["B"+str(cell.row)].value )
#         else:
#             ws_new[cell.coordinate].value = ws_previous[cell.coordinate].value
#             ws_new[cell.coordinate].font = copy(ws_previous[cell.coordinate].font)
#             ws_new[cell.coordinate].border = copy(ws_previous[cell.coordinate].border)
#             ws_new[cell.coordinate].fill = copy(ws_previous[cell.coordinate].fill)
#             ws_new[cell.coordinate].number_format = copy(ws_previous[cell.coordinate].number_format)
#             ws_new[cell.coordinate].protection = copy(ws_previous[cell.coordinate].protection)
#             ws_new[cell.coordinate].alignment = copy(ws_previous[cell.coordinate].alignment)
    
#             mark_newly_closed_cell(cell.row, ws_new["M"+str(cell.row)].value)

# Critical	Triage	Comments column
ws_new.column_dimensions[get_column_letter(28)].width = ws_previous.column_dimensions[get_column_letter(28)].width    
ws_new.column_dimensions[get_column_letter(29)].width = ws_previous.column_dimensions[get_column_letter(29)].width    
ws_new.column_dimensions[get_column_letter(30)].width = ws_previous.column_dimensions[get_column_letter(30)].width    

put_issue_triage()

# change sheet name of new comsolidated file
date = get_current_date()
fw_format = date_to_fw_format (date)
year = get_current_year()
sheetname = f"{year}_{fw_format}"
 
print("sheetname: "+sheetname) 

# position = new_file_name.find('FW')   # jinha comment FW05.1 
# print ('position', position)
# print ('slice', new_file_name[position:position+4])

# jinha comment FW05.1
ws_new.title = sheetname 

wb_new.save(new_file_name)
wb_old.save(old_file_name)
wb_new.close()
wb_old.close()

