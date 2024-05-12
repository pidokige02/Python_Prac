from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill

# gemini R3/R4/R5 의 cell 을 확인하여 match 된 cell 이 resolved, verified, closed되 었으면 true 와 같이 status 를 return 함 
def check_cloned_spr_completed(sheet, value):
    for row in sheet.iter_rows(min_row=2,min_col=2, max_col=2): # check coloned ID in osprey_link_2024FWxx..xlsx
        for cell in row:
            if(sheet[cell.coordinate].value == value):
                status = sheet["M"+str(cell.row)].value
                if((status == "Verified") or (status == "Resolved") or (status == "Closed")):
                    return True
                else:
                    return False
    return False

# Osprey r4 가 still open 되어 있으면서 genimi 가 close 되어 있으면 색으로 표시를 함
def mark_closed_spr (sheet, value):  # mark closed spr skyblue in ospret_r4
    for row in sheet.iter_rows(min_row=2,min_col=2, max_col=2):
        for cell in row: 
            if(sheet[cell.coordinate].value == value):
                status = sheet["M"+str(cell.row)].value
                if((status != "Verified") and (status != "Resolved") and (status != "Closed")):
                    sheet["B"+str(cell.row)].fill = PatternFill(fgColor="87CEEB", fill_type = "solid")
                    sheet["C"+str(cell.row)].fill = PatternFill(fgColor="87CEEB", fill_type = "solid")
                return


def  xl_clone_analysis (link_file_name, report_file_name_osprey_r4, report_file_name_for_gemini_r3, report_file_name_for_gemini_r4, report_file_name_for_gemini_r5):

    wb_link = load_workbook(link_file_name)
    wb_osprey_r4 = load_workbook(report_file_name_osprey_r4)

    wb_gemini_r3 = load_workbook(report_file_name_for_gemini_r3)
    wb_gemini_r4 = load_workbook(report_file_name_for_gemini_r4)
    wb_gemini_r5 = load_workbook(report_file_name_for_gemini_r5)

    ws_link = wb_link[wb_link.sheetnames[0]]  
    ws_osprey_r4 = wb_osprey_r4[wb_osprey_r4.sheetnames[0]] 

    ws_gemini_r3 = wb_gemini_r3[wb_gemini_r3.sheetnames[0]]  
    ws_gemini_r4 = wb_gemini_r4[wb_gemini_r4.sheetnames[0]]  
    ws_gemini_r5 = wb_gemini_r5[wb_gemini_r5.sheetnames[0]] 

    for row in ws_link.iter_rows(min_row=2,min_col=4, max_col=4): # check coloned ID in osprey_link_2024FWxx..xlsx
          for cell in row:
            if(ws_link[cell.coordinate].value):
                ret = check_cloned_spr_completed(ws_gemini_r3, ws_link[cell.coordinate].value)
                if ret == True:
                    mark_closed_spr (ws_osprey_r4, ws_link["B"+str(cell.row)].value)
                    
                ret = check_cloned_spr_completed(ws_gemini_r4, ws_link[cell.coordinate].value)
                if ret == True:
                    mark_closed_spr (ws_osprey_r4, ws_link["B"+str(cell.row)].value)

                ret = check_cloned_spr_completed(ws_gemini_r5, ws_link[cell.coordinate].value)
                if ret == True:
                    mark_closed_spr (ws_osprey_r4, ws_link["B"+str(cell.row)].value)


    wb_link.save(link_file_name)
    wb_osprey_r4.save(report_file_name_osprey_r4)
    wb_gemini_r3.save(report_file_name_for_gemini_r3)
    wb_gemini_r4.save(report_file_name_for_gemini_r4)
    wb_gemini_r5.save(report_file_name_for_gemini_r5)

    wb_link.close()
    wb_osprey_r4.close()
    wb_gemini_r3.close()
    wb_gemini_r4.close()
    wb_gemini_r5.close()
                       

